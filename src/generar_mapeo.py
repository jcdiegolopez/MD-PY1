import openpyxl
import json
import os
from difflib import SequenceMatcher

def generar_mapeo_hojas():
    """
    Lee todos los archivos Excel y crea un mapeo de cada temática
    a la hoja correspondiente en cada año.
    
    Lee el título de las tablas directamente desde la fila 6 de cada hoja.
    Guarda: mapeo_hojas.json
    Estructura: {temática: {año: nombre_hoja}}
    """
    
    # Las 16 temáticas que buscamos
    tematicas = [
        "Defunciones por año de ocurrencia, según departamento de residencia del difunto(a)",
        "Defunciones por departamento de ocurrencia, según departamento de residencia del difunto(a)",
        "Defunciones por sexo, según departamento de residencia del difunto(a) y edades simples",
        "Defunciones por sexo, según departamento de residencia del difunto(a) y grupos de edad",
        "Defunciones por sexo, según departamento de residencia del difunto(a), estado civil y grupos de edad",
        "Defunciones por sexo, según edad y causas de muerte",
        "Defunciones por sexo, según departamento de residencia del difunto(a) y causas de muerte",
        "Defunciones por tipo de certificación, según departamento y municipio de residencia del difunto(a)",
        "Defunciones por tipo de asistencia recibida, según departamento y municipio de residencia del difunto(a)",
        "Defunciones por lugar de ocurrencia, según departamento y municipio de residencia del difunto(a)",
        "Defunciones infantiles, neonatales y postneonatales por sexo, según departamento de residencia y edad",
        "Defunciones neonatales por sexo, según edad y causas de muerte",
        "Defunciones postneonatales por sexo, según edad y causas de muerte",
        "Defunciones por mes de ocurrencia,  según día de ocurrencia",
        "Defunciones por pueblo de pertenencia del difunto(a), según departamento de residencia",
        "Defunciones por causas externas y sexo, según departamento de residencia del difunto(a)"
    ]
    
    directorio_data = 'data/defunciones'
    años = ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024']
    
    print("=" * 80)
    print("GENERANDO MAPEO DE HOJAS")
    print("=" * 80)
    
    mapeo_resultado = {tematica: {} for tematica in tematicas}
    
    def similitud_texto(texto1, texto2):
        """Calcula la similitud entre dos textos (0-1)"""
        return SequenceMatcher(None, texto1.lower(), texto2.lower()).ratio()
    
    # Por cada año
    for año in años:
        print(f"\n{'='*80}")
        print(f"PROCESANDO {año}")
        print(f"{'='*80}")
        
        archivo = os.path.join(directorio_data, f'{año}.xlsx')
        
        if not os.path.exists(archivo):
            print(f"⚠ {archivo} no existe")
            continue
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            nombres_hojas = wb.sheetnames
            
            print(f"✓ Archivo: {año}.xlsx")
            
            # Leer los títulos de cada hoja (fila 6 para nuevos, fila 1 para antiguos)
            titulos_por_hoja = {}
            for nombre_hoja in nombres_hojas:
                try:
                    ws = wb[nombre_hoja]
                    # Leer ambas posibles ubicaciones
                    titulo_a6 = ws['A6'].value
                    titulo_a1 = ws['A1'].value
                    
                    # Determinar cuál es el título real basado en la longitud
                    # Los títulos suelen ser largos (>30 caracteres)
                    titulo = None
                    if titulo_a6 and len(str(titulo_a6)) > 30:
                        titulo = titulo_a6
                    elif titulo_a1 and len(str(titulo_a1)) > 30:
                        titulo = titulo_a1
                    elif titulo_a6:
                        titulo = titulo_a6
                    elif titulo_a1:
                        titulo = titulo_a1
                    
                    if titulo:
                        titulos_por_hoja[nombre_hoja] = str(titulo).strip()
                except:
                    pass
            
            wb.close()
            
            # Por cada temática
            for idx, tematica in enumerate(tematicas, 1):
                hoja_encontrada = None
                
                # Buscar por similitud de título
                mejores = []
                for nombre_hoja, titulo in titulos_por_hoja.items():
                    similitud = similitud_texto(tematica, titulo)
                    if similitud >= 0.75:  # Umbral alto
                        mejores.append((similitud, nombre_hoja))
                
                if mejores:
                    mejores.sort(reverse=True)
                    hoja_encontrada = mejores[0][1]
                
                # Si se encuentra, guardar
                if hoja_encontrada:
                    mapeo_resultado[tematica][año] = hoja_encontrada
                    print(f"    [{idx:2d}] ✓ {tematica[:60]}... → {hoja_encontrada}")
                else:
                    mapeo_resultado[tematica][año] = None
                    print(f"    [{idx:2d}] ❌ {tematica[:60]}... → NO ENCONTRADO")
        
        except Exception as e:
            print(f"❌ Error procesando {año}: {e}")
    
    # Guardar resultado
    print(f"\n{'='*80}")
    print("GUARDANDO MAPEO")
    print(f"{'='*80}")
    
    with open('data/json/mapeo_hojas.json', 'w', encoding='utf-8') as f:
        json.dump(mapeo_resultado, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ Mapeo guardado en: mapeo_hojas.json")
    
    # Resumen
    print(f"\n{'='*80}")
    print("RESUMEN")
    print(f"{'='*80}")
    
    total_encontrados = sum(
        1 for tematica in mapeo_resultado 
        for año in mapeo_resultado[tematica] 
        if mapeo_resultado[tematica][año] is not None
    )
    total_esperados = len(tematicas) * len(años)
    
    print(f"\nMapeados: {total_encontrados}/{total_esperados}")
    print(f"Éxito: {(total_encontrados/total_esperados)*100:.1f}%")
    
    # Mostrar temáticas con problemas
    sin_mapeo = {}
    for tematica in tematicas:
        años_sin_mapeo = [año for año in años if mapeo_resultado[tematica].get(año) is None]
        if años_sin_mapeo:
            sin_mapeo[tematica] = años_sin_mapeo
    
    if sin_mapeo:
        print(f"\n⚠ Temáticas con años sin mapear:")
        for tematica, años_list in sin_mapeo.items():
            print(f"  - {tematica[:60]}... → Años: {', '.join(años_list)}")
    
    print(f"\n{'='*80}")
    print("PROCESO COMPLETADO")
    print(f"{'='*80}")

if __name__ == "__main__":
    generar_mapeo_hojas()
