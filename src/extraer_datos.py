import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime

def extraer_datos_tabla(archivo_excel, nombre_hoja, estructura_tabla):
    """
    Extrae los datos de una tabla Excel basándose en su estructura.
    
    Retorna: {
        'encabezados': [...],
        'datos': [
            {...}  # cada fila como dict
        ]
    }
    """
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws = wb[nombre_hoja]
        
        # Obtener encabezados de la estructura
        encabezados_estructura = estructura_tabla['encabezados']
        
        # Convertir estructura de encabezados a lista plana de nombres de columna
        if estructura_tabla['tipo'] == 'agrupado':
            # Para tablas agrupadas, crear nombres como "grupo_subtitulo"
            nombres_columnas = []
            for titulo, valores in encabezados_estructura.items():
                if isinstance(valores, list):
                    # Tiene subtítulos
                    for subtitulo in valores:
                        nombres_columnas.append(f"{titulo}_{subtitulo}")
                else:
                    # Es un string (sin subtítulos)
                    nombres_columnas.append(titulo)
        else:
            # Para tablas simples, los nombres son directos
            nombres_columnas = encabezados_estructura
        
        # Leer datos a partir de fila 10
        datos = []
        fila_inicio = 10
        
        # Encontrar cuántas filas tiene datos
        fila_actual = fila_inicio
        max_filas = 10000  # Límite de búsqueda
        
        while fila_actual <= fila_inicio + max_filas:
            # Leer valores de la fila
            valores = []
            tiene_datos = False
            
            for col in range(1, len(nombres_columnas) + 1):
                valor = ws.cell(fila_actual, col).value
                valores.append(valor)
                if valor is not None:
                    tiene_datos = True
            
            # Si la fila no tiene datos, terminar
            if not tiene_datos:
                break
            
            # Crear diccionario de fila
            fila_dict = {}
            for col_idx, nombre_col in enumerate(nombres_columnas):
                fila_dict[nombre_col] = valores[col_idx] if col_idx < len(valores) else None
            
            datos.append(fila_dict)
            fila_actual += 1
        
        wb.close()
        
        return {
            'encabezados': nombres_columnas,
            'datos': datos,
            'total_filas': len(datos)
        }
    
    except Exception as e:
        print(f"  Error extrayendo datos: {e}")
        return {
            'encabezados': [],
            'datos': [],
            'error': str(e),
            'total_filas': 0
        }

def main():
    # Configuración
    directorio_data = 'data/defunciones'
    mapeo_json = 'data/json/mapeo_hojas.json'
    estructura_json = 'data/json/estructura_completa.json'
    datos_json = 'data/json/datos_completos.json'
    años = ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024']
    
    print("=" * 80)
    print("EXTRAYENDO DATOS DE TODAS LAS TABLAS")
    print("=" * 80)
    
    # Cargar mapeo y estructura
    if not os.path.exists(mapeo_json):
        print(f"\n❌ Error: {mapeo_json} no existe")
        return
    
    if not os.path.exists(estructura_json):
        print(f"\n❌ Error: {estructura_json} no existe")
        return
    
    print(f"\n✓ Leyendo {mapeo_json}...")
    with open(mapeo_json, 'r', encoding='utf-8') as f:
        mapeo = json.load(f)
    
    print(f"✓ Leyendo {estructura_json}...")
    with open(estructura_json, 'r', encoding='utf-8') as f:
        estructura = json.load(f)
    
    # Resultado: {tema: {año: {datos}}}
    datos_resultado = {}
    resumen_años = {}
    
    # Por cada año
    for año in años:
        print(f"\n{'='*80}")
        print(f"PROCESANDO AÑO {año}")
        print(f"{'='*80}")
        
        archivo_año = os.path.join(directorio_data, f'{año}.xlsx')
        
        if not os.path.exists(archivo_año):
            print(f"⚠ {archivo_año} no existe, saltando...")
            resumen_años[año] = {'procesadas': 0, 'error': 'Archivo no existe'}
            continue
        
        print(f"✓ Procesando {año}.xlsx...")
        
        temáticas_procesadas = 0
        total_filas_año = 0
        
        # Por cada temática
        for tematica, años_map in mapeo.items():
            nombre_hoja = años_map.get(año)
            
            if not nombre_hoja:
                continue
            
            # Obtener estructura de esta temática y año
            if tematica not in estructura or año not in estructura[tematica]:
                print(f"  ⚠ No hay estructura para {tematica[:50]}... en {año}")
                continue
            
            estructura_tabla = estructura[tematica][año]
            
            # Inicializar entrada para temática si no existe
            if tematica not in datos_resultado:
                datos_resultado[tematica] = {}
            
            # Extraer datos
            datos = extraer_datos_tabla(archivo_año, nombre_hoja, estructura_tabla)
            
            datos_resultado[tematica][año] = {
                'encabezados': datos['encabezados'],
                'datos': datos['datos'],
                'total_filas': datos['total_filas'],
                'tipo_tabla': estructura_tabla['tipo'],
                'hoja': nombre_hoja
            }
            
            temáticas_procesadas += 1
            total_filas_año += datos['total_filas']
            
            print(f"  [{temáticas_procesadas:2d}] {tematica[:50]}... ({datos['total_filas']:6d} filas)")
        
        resumen_años[año] = {
            'procesadas': temáticas_procesadas,
            'total_filas': total_filas_año
        }
    
    # Guardar resultado
    print(f"\n{'='*80}")
    print("GUARDANDO RESULTADO")
    print(f"{'='*80}")
    
    with open(datos_json, 'w', encoding='utf-8') as f:
        json.dump(datos_resultado, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ Datos guardados en: {datos_json}")
    
    # Resumen
    print(f"\n{'='*80}")
    print("RESUMEN")
    print(f"{'='*80}")
    
    print(f"\nAños procesados: {len(años)}")
    for año in años:
        if año in resumen_años:
            info = resumen_años[año]
            if 'error' in info:
                print(f"  {año}: ⚠ {info['error']}")
            else:
                print(f"  {año}: {info['procesadas']} temáticas, {info['total_filas']:,} filas totales")
    
    total_temáticas = len(datos_resultado)
    total_filas_general = sum(
        len(datos_resultado[t][a]['datos'])
        for t in datos_resultado
        for a in datos_resultado[t]
    )
    
    print(f"\nTotal temáticas: {total_temáticas}")
    print(f"Total filas (todas las tablas): {total_filas_general:,}")
    
    print(f"\n{'='*80}")
    print("PROCESO COMPLETADO")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
