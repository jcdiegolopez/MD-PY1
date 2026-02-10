import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
from pathlib import Path

def detectar_tipo_encabezado(ws, año=None):
    """
    Detecta si una tabla tiene encabezados simples (1 fila) o agrupados (2 filas).
    
    Para 2022 y anteriores: revisa filas 1-2
    Para 2023 y posteriores: revisa filas 8-9
    
    Lógica:
    - AGRUPADA: Fila N+1 tiene sub-encabezados donde fila N está vacío (estructura multi-nivel)
    - SIMPLE: Fila N tiene encabezados, fila N+1 tiene DATOS (números, texto de datos)
    
    Retorna: 'simple' o 'agrupado'
    """
    # Determinar filas según año
    if año and int(año) <= 2022:
        fila_principal = 1
        fila_secundaria = 2
    else:
        fila_principal = 8
        fila_secundaria = 9
    
    # Contar patrones
    fila_principal_con_valor = 0
    fila_secundaria_con_valor = 0
    fila_secundaria_donde_principal_vacio = 0  # Columnas donde principal es vacío pero secundaria tiene valor (sub-encabezados)
    fila_secundaria_donde_principal_lleno = 0   # Columnas donde ambas tienen valor
    
    for col in range(1, 100):
        col_letter = get_column_letter(col)
        valor_principal = ws[f'{col_letter}{fila_principal}'].value
        valor_secundaria = ws[f'{col_letter}{fila_secundaria}'].value
        
        if valor_principal:
            fila_principal_con_valor += 1
        
        if valor_secundaria:
            fila_secundaria_con_valor += 1
            
            if not valor_principal:
                # Fila secundaria tiene valor pero principal no -> es sub-encabezado
                fila_secundaria_donde_principal_vacio += 1
            else:
                # Ambas filas tienen valor -> probablemente es dato
                fila_secundaria_donde_principal_lleno += 1
        
        # Parar si llegamos a muchas columnas vacías
        if col > 50 and fila_principal_con_valor == 0 and fila_secundaria_con_valor == 0:
            break
    
    # Decisión mejorada:
    # Si hay SUB-encabezados, es agrupado
    # BUT: Si la mayoría de fila_secundaria contiene valores donde fila_principal TAMBIÉN tiene valores,
    #      probablemente sea datos simples (no sub-encabezados)
    
    if fila_secundaria_donde_principal_vacio > 0 and fila_secundaria_donde_principal_vacio >= fila_secundaria_donde_principal_lleno:
        # Hay sub-encabezados y son más o igual que datos
        return 'agrupado'
    else:
        # Fila secundaria probablemente tiene datos, no sub-encabezados
        return 'simple'

def leer_encabezados_simple(ws, max_col=50, año=None):
    """
    Lee encabezados de una tabla simple (una sola fila).
    
    Para 2022 y anteriores: lee encabezados de fila 1
    Para 2023 y posteriores: lee encabezados de fila 8
    """
    # Determinar fila según año
    if año and int(año) <= 2022:
        fila = 1
    else:
        fila = 8
    
    encabezados = []
    
    col = 1
    while col <= max_col:
        col_letter = get_column_letter(col)
        valor = ws[f'{col_letter}{fila}'].value
        
        if not valor:
            break
        
        nombre_col = str(valor).strip()
        encabezados.append(nombre_col)
        col += 1
    
    return encabezados

def leer_encabezados_agrupado(ws, max_col=50, año=None):
    """
    Lee encabezados de una tabla agrupada (dos filas).
    
    Para 2022 y anteriores: encabezados principales en fila 1, sub-encabezados en fila 2
    Para 2023 y posteriores: encabezados principales en fila 8, sub-encabezados en fila 9
    
    Detecta celdas combinadas.
    
    Retorna una estructura armada con título principal y subtítulos:
    - Si tiene subtítulos reales: {"titulo_principal": ["subtitulo1", "subtitulo2"]}
    - Si NO tiene subtítulos: {"titulo_principal": "solo_nombre"}
    """
    # Determinar filas según año
    if año and int(año) <= 2022:
        fila_principal = 1
        fila_secundaria = 2
    else:
        fila_principal = 8
        fila_secundaria = 9
    
    estructura = {}
    grupos_con_subtitulos = {}  # Para rastrear si un grupo tiene subtítulos reales
    
    # Mapear celdas combinadas
    mapa_combinadas = {}
    for rango_combinado in ws.merged_cells.ranges:
        min_col = rango_combinado.min_col
        max_col_rango = rango_combinado.max_col
        min_row = rango_combinado.min_row
        max_row = rango_combinado.max_row
        
        # Si es una celda combinada horizontal en fila principal
        if min_row == fila_principal and max_row == fila_principal and min_col < max_col_rango:
            col_letter = get_column_letter(min_col)
            valor = ws[f'{col_letter}{fila_principal}'].value
            
            for col in range(min_col, max_col_rango + 1):
                mapa_combinadas[col] = valor
    
    # Leer encabezados
    col = 1
    while col <= max_col:
        col_letter = get_column_letter(col)
        
        valor_principal = ws[f'{col_letter}{fila_principal}'].value
        valor_secundario = ws[f'{col_letter}{fila_secundaria}'].value
        
        # Aplicar mapeo de combinadas
        if col in mapa_combinadas:
            valor_principal = mapa_combinadas[col]
        
        # Si no hay valores, terminar
        if not valor_principal and not valor_secundario:
            break
        
        # Procesar según el valor principal
        if valor_principal:
            valor_principal = str(valor_principal).strip()
            
            # Inicializar grupo si no existe
            if valor_principal not in estructura:
                estructura[valor_principal] = []
                grupos_con_subtitulos[valor_principal] = False
        
        # Agregar subtítulo si existe
        if valor_secundario:
            valor_secundario = str(valor_secundario).strip()
            if valor_principal:
                estructura[valor_principal].append(valor_secundario)
                grupos_con_subtitulos[valor_principal] = True
            else:
                # No debería ocurrir en agrupadas, pero por seguridad
                if "sin_grupo" not in estructura:
                    estructura["sin_grupo"] = []
                    grupos_con_subtitulos["sin_grupo"] = True
                estructura["sin_grupo"].append(valor_secundario)
        
        col += 1
    
    # Convertir a strings aquellos grupos sin subtítulos reales
    resultado = {}
    for titulo, valores in estructura.items():
        if grupos_con_subtitulos.get(titulo, False) and valores:
            # Tiene subtítulos reales, mantener como lista
            resultado[titulo] = valores
        else:
            # No tiene subtítulos reales, convertir a string
            resultado[titulo] = titulo
    
    return resultado

def extraer_estructura_tabla(archivo_excel, nombre_hoja, año=None):
    """
    Extrae solo la estructura (encabezados) de una tabla.
    Detecta automáticamente si es simple o agrupada.
    
    Para 2022 y anteriores: lee de filas 1-2
    Para 2023 y posteriores: lee de filas 8-9
    
    Retorna para SIMPLE: 
        {'tipo': 'simple', 'encabezados': [...], 'total_columnas': int}
    
    Retorna para AGRUPADO:
        {'tipo': 'agrupado', 'encabezados': {titulo: [subtitulos] o titulo: "nombre"}, 'total_columnas': int}
    """
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws = wb[nombre_hoja]
        
        # Detectar tipo de encabezado
        tipo = detectar_tipo_encabezado(ws, año=año)
        
        # Leer encabezados según tipo
        if tipo == 'agrupado':
            encabezados = leer_encabezados_agrupado(ws, año=año)
            # Contar total de columnas (suma de subtítulos + strings individuales)
            total_cols = 0
            for valor in encabezados.values():
                if isinstance(valor, list):
                    total_cols += len(valor)
                else:  # Es un string sin subtítulos
                    total_cols += 1
        else:
            encabezados = leer_encabezados_simple(ws, año=año)
            total_cols = len(encabezados)
        
        wb.close()
        
        return {
            'tipo': tipo,
            'encabezados': encabezados,
            'total_columnas': total_cols
        }
    
    except Exception as e:
        print(f"  Error: {e}")
        return {'tipo': 'error', 'encabezados': {}, 'error': str(e), 'total_columnas': 0}

def main():
    # Configuración
    directorio_data = 'data/defunciones'
    mapeo_json = 'data/json/mapeo_hojas.json'
    estructura_json = 'data/json/estructura_completa.json'
    años_procesamiento = ['2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024']
    
    print("=" * 70)
    print("EXTRAYENDO ESTRUCTURA DE COLUMNAS - TODOS LOS AÑOS (2015-2024)")
    print("=" * 70)
    
    # Verificar si existe mapeo_hojas.json
    if not os.path.exists(mapeo_json):
        print(f"\n❌ Error: {mapeo_json} no existe. Ejecuta primero el script de mapeo.")
        return
    
    # Cargar mapeo
    print(f"\n✓ Leyendo {mapeo_json}...")
    with open(mapeo_json, 'r', encoding='utf-8') as f:
        mapeo = json.load(f)
    
    # Estructura resultado: {tema: {año: estructura}}
    estructura_resultado = {}
    resumen_años = {}
    
    # Por cada año
    for año in años_procesamiento:
        print(f"\n{'='*70}")
        print(f"PROCESANDO AÑO {año}")
        print(f"{'='*70}")
        
        archivo_año = os.path.join(directorio_data, f'{año}.xlsx')
        
        if not os.path.exists(archivo_año):
            print(f"\n⚠ {archivo_año} no existe, saltando...")
            resumen_años[año] = {'procesadas': 0, 'error': 'Archivo no existe'}
            continue
        
        print(f"✓ Procesando {año}.xlsx...")
        
        temáticas_procesadas = 0
        agrupadas_año = 0
        simples_año = 0
        total_cols_año = 0
        
        # Por cada temática
        for tematica, años_map in mapeo.items():
            nombre_hoja = años_map.get(año)
            
            if not nombre_hoja:
                continue
            
            # Inicializar entrada para temática si no existe
            if tematica not in estructura_resultado:
                estructura_resultado[tematica] = {}
            
            # Extraer estructura (pasar el año para detectar la correcta ubicación de encabezados)
            estructura = extraer_estructura_tabla(archivo_año, nombre_hoja, año=año)
            
            estructura['hoja'] = nombre_hoja
            estructura_resultado[tematica][año] = estructura
            
            temáticas_procesadas += 1
            total_cols_año += estructura['total_columnas']
            
            if estructura['tipo'] == 'agrupado':
                agrupadas_año += 1
            else:
                simples_año += 1
            
            print(f"  [{temáticas_procesadas}] {tematica[:50]}... ({estructura['tipo']}, {estructura['total_columnas']} cols)")
        
        resumen_años[año] = {
            'procesadas': temáticas_procesadas,
            'agrupadas': agrupadas_año,
            'simples': simples_año,
            'total_columnas': total_cols_año
        }
    
    # Guardar resultado
    print("\n" + "=" * 70)
    print("GUARDANDO RESULTADO")
    print("=" * 70)
    
    with open(estructura_json, 'w', encoding='utf-8') as f:
        json.dump(estructura_resultado, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ Estructura guardada en: {estructura_json}")
    
    # Resumen general
    print("\n" + "=" * 70)
    print("RESUMEN GENERAL")
    print("=" * 70)
    
    print(f"\nAños procesados: {len(años_procesamiento)}")
    for año in años_procesamiento:
        if año in resumen_años:
            info = resumen_años[año]
            if 'error' in info:
                print(f"  {año}: ⚠ {info['error']}")
            else:
                print(f"  {año}: {info['procesadas']} temáticas ({info['agrupadas']} agrupadas, {info['simples']} simples, {info['total_columnas']} cols)")
    
    total_temáticas = len(estructura_resultado)
    total_años_completos = sum(1 for y in años_procesamiento if y in resumen_años and 'error' not in resumen_años[y])
    
    print(f"\nTotal temáticas: {total_temáticas}")
    print(f"Años completos: {total_años_completos}/{len(años_procesamiento)}")
    
    print("\n" + "=" * 70)
    print("PROCESO COMPLETADO")
    print("=" * 70)

def leer_estructura_columnas(archivo_excel, nombre_hoja):
    """
    Esta función está deprecada. Usar extraer_estructura_tabla() en su lugar.
    """
    pass

if __name__ == "__main__":
    main()
