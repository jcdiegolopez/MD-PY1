import openpyxl

# Probar con un archivo de 2020
archivo = 'data/defunciones/2020.xlsx'
wb = openpyxl.load_workbook(archivo, data_only=True)

print("=" * 80)
print("ANÁLISIS DE TODAS LAS HOJAS EN 2020.xlsx")
print("=" * 80)

for nombre_hoja in wb.sheetnames:
    print(f"\n📄 Hoja: {nombre_hoja}")
    ws = wb[nombre_hoja]
    
    # Revisar A1 y A6
    a1 = ws['A1'].value
    a6 = ws['A6'].value
    
    if a1:
        print(f"  A1: {str(a1)[:80]}")
    if a6:
        print(f"  A6: {str(a6)[:80]}")

wb.close()
